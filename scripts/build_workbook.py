"""6-tab Excel builder for the Competitor Analyzer plugin.

Tabs:
    1. Summary           — biggest gap per section (primary vs best competitor)
    2. Off-Page          — Ahrefs metrics
    3. On-Page           — pages, blog, footer, homepage, case studies
    4. Technical         — metadata + Core Web Vitals + rendering type
    5. Hygiene           — breadcrumbs + EEAT scores + slug audit
    6. GEO               — schema presence + AI Overview readiness scores

REDESIGN vs the Suite's export.py — the source had cell-alignment bugs,
wrap-text was inconsistent, blank cells from missing-key lookups, etc.
This redesign:
  - **rows = metrics, columns = domains** (primary first)
  - every cell uses `_val()` which converts None/"" → "—" (em-dash) so
    nothing is ever truly blank
  - `wrap_text=True` on every data cell, with explicit row-height tuning
    proportional to line count
  - frozen header row + frozen first column (the metric labels)
  - tier-colored summary cells (the largest gaps highlighted)

Scorecard JSON contract (what Claude writes):
{
    "run_date": "YYYY-MM-DD",
    "sections_enabled": ["off","on","tech","hyg","geo"],
    "domains": [
        {
            "role": "primary"|"competitor",
            "domain": str,
            "off_page": {...},                # RAW from prepare
            "on_page": {...},                 # RAW from prepare
            "technical": {...},               # RAW from prepare
            "hygiene": {...},                 # RAW from prepare
            "geo": {...},                     # RAW from prepare
            "classifications": {              # NEW — Claude fills these
                "blog_categories": [...],
                "footer_tags": {"Trust": [...], "Resources": [...], ...},
                "homepage_clarity": {"score": 1-5, "verdict": str, "strengths": [...], "weaknesses": [...]},
                "case_studies_verified": [{url, title}, ...],
                "eeat_scores": {"overall": 1-5, "experience": 1-5, "expertise": 1-5, "authoritativeness": 1-5, "trustworthiness": 1-5, "summary": str},
                "slug_audit": {"flagged_count": int, "issues": [...]},
                "geo_scores": {"answer_first": 1-5, "llm_citation_readiness": 1-5, "entity_coverage": 1-5, "notes": str}
            }
        },
        ...
    ],
    "summary": {
        "biggest_gaps": [
            {"section": "On-Page", "metric": "Homepage clarity", "primary_value": "2", "best_competitor_value": "4", "best_competitor": "X.com"},
            ...
        ]
    }
}
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styling ──────────────────────────────────────────────────────────
HEADER_FILL = "FF367588"
ALT_ROW_FILL = "FFF7FAFA"
PRIMARY_FILL = "FFEAF4F4"        # primary-domain column shading
GAP_HIGHLIGHT_FILL = "FFFFE082"  # yellow — for biggest-gap rows on Summary tab

WHITE = "FFFFFFFF"
DARK = "FF1D1D20"

ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color=WHITE)
DATA_FONT = Font(name=ARIAL, size=9, color=DARK)
BOLD_FONT = Font(name=ARIAL, size=9, bold=True, color=DARK)
LARGE_BOLD = Font(name=ARIAL, size=12, bold=True, color=DARK)

THIN = Side(border_style="thin", color="FFCFCFCF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_TOP_WRAP = Alignment(vertical="top", horizontal="left", wrap_text=True)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _val(v: Any) -> str:
    """Defensive conversion. None / empty / missing → em-dash. Lists → '\n'-joined."""
    if v is None or v == "" or v == []:
        return "—"
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, list):
        clean = [str(x) for x in v if x not in (None, "", [])]
        return "\n".join(clean) if clean else "—"
    if isinstance(v, dict):
        if not v:
            return "—"
        return "\n".join(f"{k}: {_val(val)}" for k, val in v.items())
    s = str(v).strip()
    return s if s else "—"


def _h(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    c.alignment = ALIGN_CENTER
    c.border = BORDER


def _d(ws, row: int, col: int, value: Any, *,
       font: Font = DATA_FONT,
       fill: str | None = None,
       align: Alignment = ALIGN_TOP_WRAP) -> None:
    c = ws.cell(row=row, column=col, value=_val(value))
    c.font = font
    c.alignment = align
    c.border = BORDER
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _set_row_heights(ws, start_row: int, end_row: int, base: float = 40) -> None:
    """Set row heights proportional to multi-line content."""
    for r in range(start_row, end_row + 1):
        max_lines = 1
        for col_cells in ws.iter_cols(min_row=r, max_row=r, values_only=True):
            for v in col_cells:
                if isinstance(v, str):
                    max_lines = max(max_lines, v.count("\n") + 1)
        ws.row_dimensions[r].height = max(base, max_lines * 15 + 10)


def _domain_label(d: dict) -> str:
    role = "(primary)" if d.get("role") == "primary" else ""
    return f"{d.get('domain', '?')}\n{role}".strip()


# ── Section: Off-Page ────────────────────────────────────────────────

def _build_off_page(ws, scorecard: dict) -> None:
    ws.title = "Off-Page"
    domains = scorecard.get("domains", [])
    headers = ["Metric"] + [_domain_label(d) for d in domains]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    rows = [
        ("Domain Rating (DR)", lambda d: d.get("off_page", {}).get("dr")),
        ("Backlinks (now)", lambda d: d.get("off_page", {}).get("backlinks_now")),
        ("Backlinks (90d ago)", lambda d: d.get("off_page", {}).get("backlinks_90d")),
        ("Backlinks Δ (90d)", lambda d: d.get("off_page", {}).get("backlinks_delta")),
        ("Referring Domains (now)", lambda d: d.get("off_page", {}).get("refdomains_now")),
        ("Referring Domains (90d ago)", lambda d: d.get("off_page", {}).get("refdomains_90d")),
        ("Referring Domains Δ (90d)", lambda d: d.get("off_page", {}).get("refdomains_delta")),
        ("Organic Traffic (est)", lambda d: d.get("off_page", {}).get("organic_traffic")),
        ("Top Anchors", lambda d: d.get("off_page", {}).get("top_anchors")),
        ("Status", lambda d: "Available" if d.get("off_page", {}).get("available") else f"N/A: {d.get('off_page', {}).get('reason', 'unknown')}"),
    ]
    for r, (label, getter) in enumerate(rows, start=2):
        _d(ws, r, 1, label, font=BOLD_FONT)
        for col, d in enumerate(domains, start=2):
            fill = PRIMARY_FILL if d.get("role") == "primary" else None
            _d(ws, r, col, getter(d), fill=fill, align=ALIGN_CENTER)
    _set_col_widths(ws, [30] + [28] * len(domains))
    _set_row_heights(ws, 2, len(rows) + 1)
    ws.freeze_panes = "B2"


# ── Section: On-Page ─────────────────────────────────────────────────

def _build_on_page(ws, scorecard: dict) -> None:
    ws.title = "On-Page"
    domains = scorecard.get("domains", [])
    headers = ["Metric"] + [_domain_label(d) for d in domains]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    def _classifications(d):
        return d.get("classifications", {})

    rows = [
        ("Total indexed pages", lambda d: d.get("on_page", {}).get("pages", {}).get("total_urls")),
        ("Sitemap available", lambda d: d.get("on_page", {}).get("pages", {}).get("has_sitemap")),
        ("Blog post count", lambda d: d.get("on_page", {}).get("blog", {}).get("post_count")),
        ("Blog categories (Claude clustered)", lambda d: _classifications(d).get("blog_categories")),
        ("Top 5 blog samples", lambda d: d.get("on_page", {}).get("blog", {}).get("sample_urls")),
        ("Total footer links", lambda d: d.get("on_page", {}).get("footer", {}).get("total_links")),
        ("Footer — Trust links", lambda d: _classifications(d).get("footer_tags", {}).get("Trust")),
        ("Footer — Resources links", lambda d: _classifications(d).get("footer_tags", {}).get("Resources")),
        ("Footer — Service links", lambda d: _classifications(d).get("footer_tags", {}).get("Service")),
        ("Footer — Compliance links", lambda d: _classifications(d).get("footer_tags", {}).get("Compliance")),
        ("Footer — Other links", lambda d: _classifications(d).get("footer_tags", {}).get("Other")),
        ("Homepage clarity score (1-5)", lambda d: _classifications(d).get("homepage_clarity", {}).get("score")),
        ("Homepage clarity verdict", lambda d: _classifications(d).get("homepage_clarity", {}).get("verdict")),
        ("Homepage strengths", lambda d: _classifications(d).get("homepage_clarity", {}).get("strengths")),
        ("Homepage weaknesses", lambda d: _classifications(d).get("homepage_clarity", {}).get("weaknesses")),
        ("Case studies (candidates)", lambda d: d.get("on_page", {}).get("case_studies", {}).get("total_candidates")),
        ("Case studies (Claude verified)", lambda d: len(_classifications(d).get("case_studies_verified", []) or [])),
        ("Case study sample URLs", lambda d: [c.get("url") for c in (_classifications(d).get("case_studies_verified") or [])[:5]]),
    ]
    for r, (label, getter) in enumerate(rows, start=2):
        _d(ws, r, 1, label, font=BOLD_FONT)
        for col, d in enumerate(domains, start=2):
            fill = PRIMARY_FILL if d.get("role") == "primary" else None
            _d(ws, r, col, getter(d), fill=fill)
    _set_col_widths(ws, [34] + [32] * len(domains))
    _set_row_heights(ws, 2, len(rows) + 1)
    ws.freeze_panes = "B2"


# ── Section: Technical ───────────────────────────────────────────────

def _build_technical(ws, scorecard: dict) -> None:
    ws.title = "Technical"
    domains = scorecard.get("domains", [])
    headers = ["Metric"] + [_domain_label(d) for d in domains]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    def _meta(d, key): return d.get("technical", {}).get("metadata", {}).get(key)
    def _cwv(d, strat, key): return d.get("technical", {}).get("core_web_vitals", {}).get(strat, {}).get(key)
    def _render(d, key): return d.get("technical", {}).get("rendering", {}).get(key)

    rows = [
        ("Title tag", lambda d: _meta(d, "title")),
        ("Title length", lambda d: _meta(d, "title_len")),
        ("Meta description", lambda d: _meta(d, "description")),
        ("Meta description length", lambda d: _meta(d, "desc_len")),
        ("OG image present", lambda d: _meta(d, "og_image")),
        ("Canonical URL", lambda d: _meta(d, "canonical")),
        ("Robots meta", lambda d: _meta(d, "robots")),
        ("Final URL (after redirects)", lambda d: _meta(d, "final_url")),
        ("Rendering type", lambda d: _render(d, "rendering_type")),
        ("SPA framework (if any)", lambda d: _render(d, "spa_framework")),
        ("Mobile performance score", lambda d: _cwv(d, "mobile", "performance_score")),
        ("Mobile LCP", lambda d: _cwv(d, "mobile", "lcp")),
        ("Mobile CLS", lambda d: _cwv(d, "mobile", "cls")),
        ("Mobile TBT", lambda d: _cwv(d, "mobile", "fid")),
        ("Desktop performance score", lambda d: _cwv(d, "desktop", "performance_score")),
        ("Desktop LCP", lambda d: _cwv(d, "desktop", "lcp")),
        ("Desktop CLS", lambda d: _cwv(d, "desktop", "cls")),
        ("Desktop TBT", lambda d: _cwv(d, "desktop", "fid")),
    ]
    for r, (label, getter) in enumerate(rows, start=2):
        _d(ws, r, 1, label, font=BOLD_FONT)
        for col, d in enumerate(domains, start=2):
            fill = PRIMARY_FILL if d.get("role") == "primary" else None
            _d(ws, r, col, getter(d), fill=fill)
    _set_col_widths(ws, [34] + [32] * len(domains))
    _set_row_heights(ws, 2, len(rows) + 1)
    ws.freeze_panes = "B2"


# ── Section: Hygiene ─────────────────────────────────────────────────

def _build_hygiene(ws, scorecard: dict) -> None:
    ws.title = "Hygiene"
    domains = scorecard.get("domains", [])
    headers = ["Metric"] + [_domain_label(d) for d in domains]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    def _bc(d, k): return d.get("hygiene", {}).get("breadcrumbs", {}).get(k)
    def _eeat(d, k): return d.get("classifications", {}).get("eeat_scores", {}).get(k)
    def _eeat_signal(d, page, k):
        return d.get("hygiene", {}).get("eeat_signals", {}).get(page, {}).get(k)

    rows = [
        ("Breadcrumbs present", lambda d: _bc(d, "breadcrumbs_present")),
        ("Breadcrumbs JSON-LD", lambda d: _bc(d, "breadcrumbs_jsonld")),
        ("Breadcrumbs in nav", lambda d: _bc(d, "breadcrumbs_nav")),
        ("EEAT — Overall (1-5)", lambda d: _eeat(d, "overall")),
        ("EEAT — Experience (1-5)", lambda d: _eeat(d, "experience")),
        ("EEAT — Expertise (1-5)", lambda d: _eeat(d, "expertise")),
        ("EEAT — Authoritativeness (1-5)", lambda d: _eeat(d, "authoritativeness")),
        ("EEAT — Trustworthiness (1-5)", lambda d: _eeat(d, "trustworthiness")),
        ("EEAT — Summary", lambda d: _eeat(d, "summary")),
        ("About page URL", lambda d: _eeat_signal(d, "about_page", "candidate_url")),
        ("Team page URL", lambda d: _eeat_signal(d, "team_page", "candidate_url")),
        ("Contact page URL", lambda d: _eeat_signal(d, "contact_page", "candidate_url")),
        ("Awards page URL", lambda d: _eeat_signal(d, "awards_page", "candidate_url")),
        ("Author signals found", lambda d: d.get("hygiene", {}).get("eeat_signals", {}).get("authors", {}).get("has_any_author_signals")),
        ("Slug audit — flagged count", lambda d: d.get("classifications", {}).get("slug_audit", {}).get("flagged_count")),
        ("Slug audit — top issues", lambda d: d.get("classifications", {}).get("slug_audit", {}).get("issues")),
    ]
    for r, (label, getter) in enumerate(rows, start=2):
        _d(ws, r, 1, label, font=BOLD_FONT)
        for col, d in enumerate(domains, start=2):
            fill = PRIMARY_FILL if d.get("role") == "primary" else None
            _d(ws, r, col, getter(d), fill=fill)
    _set_col_widths(ws, [34] + [32] * len(domains))
    _set_row_heights(ws, 2, len(rows) + 1)
    ws.freeze_panes = "B2"


# ── Section: GEO ─────────────────────────────────────────────────────

def _build_geo(ws, scorecard: dict) -> None:
    ws.title = "GEO"
    domains = scorecard.get("domains", [])
    headers = ["Metric"] + [_domain_label(d) for d in domains]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    def _home(d, k): return d.get("geo", {}).get("homepage", {}).get(k)
    def _atomic(d, k): return d.get("geo", {}).get("homepage", {}).get("atomic_facts_signals", {}).get(k)
    def _geo_score(d, k): return d.get("classifications", {}).get("geo_scores", {}).get(k)

    rows = [
        ("Schema markup — types present", lambda d: list((_home(d, "schemas_present") or {}).keys())),
        ("Schema markup — count of GEO-relevant types", lambda d: _home(d, "schemas_relevant_count")),
        ("FAQ Q&A pairs detected (homepage)", lambda d: _home(d, "faq_pair_count")),
        ("Specific numbers / data points (homepage)", lambda d: _atomic(d, "specific_numbers_count")),
        ("Year mentions (homepage)", lambda d: _atomic(d, "year_mentions_count")),
        ("Source-phrase count ('according to', etc.)", lambda d: _atomic(d, "source_phrase_count")),
        ("External links (homepage)", lambda d: _atomic(d, "external_links_count")),
        ("Answer-first score (1-5)", lambda d: _geo_score(d, "answer_first")),
        ("LLM Citation Readiness (1-5)", lambda d: _geo_score(d, "llm_citation_readiness")),
        ("Entity coverage (1-5)", lambda d: _geo_score(d, "entity_coverage")),
        ("GEO notes (Claude assessment)", lambda d: _geo_score(d, "notes")),
    ]
    for r, (label, getter) in enumerate(rows, start=2):
        _d(ws, r, 1, label, font=BOLD_FONT)
        for col, d in enumerate(domains, start=2):
            fill = PRIMARY_FILL if d.get("role") == "primary" else None
            _d(ws, r, col, getter(d), fill=fill)
    _set_col_widths(ws, [38] + [32] * len(domains))
    _set_row_heights(ws, 2, len(rows) + 1)
    ws.freeze_panes = "B2"


# ── Summary tab ──────────────────────────────────────────────────────

def _build_summary(ws, scorecard: dict) -> None:
    ws.title = "Summary"
    _set_col_widths(ws, [22, 28, 28, 22, 28, 50])

    c = ws.cell(row=1, column=1, value="Competitor Analyzer — Summary of Biggest Gaps")
    c.font = LARGE_BOLD
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Run context
    ws.cell(row=2, column=1, value="Run date:").font = BOLD_FONT
    ws.cell(row=2, column=2, value=scorecard.get("run_date", "—")).font = DATA_FONT
    ws.cell(row=3, column=1, value="Country:").font = BOLD_FONT
    ws.cell(row=3, column=2, value=scorecard.get("country", "—")).font = DATA_FONT
    ws.cell(row=4, column=1, value="Sections:").font = BOLD_FONT
    ws.cell(row=4, column=2, value=", ".join(scorecard.get("sections_enabled", []) or [])).font = DATA_FONT

    domains = scorecard.get("domains", [])
    ws.cell(row=5, column=1, value="Domains:").font = BOLD_FONT
    ws.cell(row=5, column=2, value="\n".join(
        f"{'(P) ' if d.get('role') == 'primary' else '     '}{d.get('domain', '')}"
        for d in domains
    )).font = DATA_FONT
    ws.cell(row=5, column=2).alignment = ALIGN_TOP_WRAP
    ws.row_dimensions[5].height = max(40, len(domains) * 18)

    # Biggest gaps header
    header_row = 7
    headers = ["Section", "Metric", "Primary value", "Best competitor", "Best competitor value", "Notes / suggested action"]
    for col, h in enumerate(headers, 1):
        _h(ws, header_row, col, h)

    gaps = scorecard.get("summary", {}).get("biggest_gaps", []) or []
    if not gaps:
        # Auto-compute a simple gaps list if Claude didn't provide one
        gaps = _auto_summary_gaps(scorecard)

    for r, gap in enumerate(gaps, start=header_row + 1):
        _d(ws, r, 1, gap.get("section"), font=BOLD_FONT, fill=GAP_HIGHLIGHT_FILL)
        _d(ws, r, 2, gap.get("metric"), font=BOLD_FONT)
        _d(ws, r, 3, gap.get("primary_value"))
        _d(ws, r, 4, gap.get("best_competitor"))
        _d(ws, r, 5, gap.get("best_competitor_value"))
        _d(ws, r, 6, gap.get("note"))

    _set_row_heights(ws, header_row + 1, header_row + len(gaps))
    ws.freeze_panes = "A8"


def _auto_summary_gaps(scorecard: dict) -> list[dict]:
    """Fallback if Claude didn't populate scorecard.summary.biggest_gaps.
    Picks a few simple numeric metrics where the primary trails."""
    gaps: list[dict] = []
    domains = scorecard.get("domains", [])
    if not domains:
        return gaps
    primary = next((d for d in domains if d.get("role") == "primary"), None)
    competitors = [d for d in domains if d.get("role") != "primary"]
    if not primary or not competitors:
        return gaps

    def _safe_num(v):
        try:
            return float(v) if v not in (None, "", "—") else None
        except (TypeError, ValueError):
            return None

    checks = [
        ("Off-Page", "Domain Rating (DR)",
         lambda d: _safe_num(d.get("off_page", {}).get("dr"))),
        ("Off-Page", "Referring Domains (now)",
         lambda d: _safe_num(d.get("off_page", {}).get("refdomains_now"))),
        ("Technical", "Mobile performance score",
         lambda d: _safe_num(d.get("technical", {}).get("core_web_vitals", {}).get("mobile", {}).get("performance_score"))),
        ("On-Page", "Blog post count",
         lambda d: _safe_num(d.get("on_page", {}).get("blog", {}).get("post_count"))),
        ("GEO", "Schemas — count of GEO-relevant types",
         lambda d: _safe_num(d.get("geo", {}).get("homepage", {}).get("schemas_relevant_count"))),
    ]
    for section, metric, getter in checks:
        p_val = getter(primary)
        comp_vals = [(c, getter(c)) for c in competitors]
        comp_vals = [(c, v) for c, v in comp_vals if v is not None]
        if not comp_vals or p_val is None:
            continue
        best_comp, best_val = max(comp_vals, key=lambda t: t[1])
        if best_val > p_val:
            gaps.append({
                "section": section,
                "metric": metric,
                "primary_value": p_val,
                "best_competitor": best_comp.get("domain"),
                "best_competitor_value": best_val,
                "note": f"Primary trails by {best_val - p_val:.1f}. Consider investing here.",
            })
    return gaps


# ── Public entry point ──────────────────────────────────────────────

_SECTION_BUILDERS = {
    "off": _build_off_page,
    "on": _build_on_page,
    "tech": _build_technical,
    "hyg": _build_hygiene,
    "geo": _build_geo,
}


def build(scorecard: dict, out_path: str) -> None:
    """Write the 6-tab Excel."""
    enabled = scorecard.get("sections_enabled") or ["off", "on", "tech", "hyg", "geo"]

    wb = Workbook()
    _build_summary(wb.active, scorecard)

    for key in ("off", "on", "tech", "hyg", "geo"):
        if key not in enabled:
            continue
        ws = wb.create_sheet()
        _SECTION_BUILDERS[key](ws, scorecard)

    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the 6-tab Competitor Analyzer Excel.")
    parser.add_argument("scorecard_json", help="Path to the scorecard JSON.")
    parser.add_argument("output_xlsx", help="Path to write the Excel workbook.")
    args = parser.parse_args()

    in_path = Path(args.scorecard_json).expanduser().resolve()
    out_path = Path(args.output_xlsx).expanduser().resolve()
    if not in_path.exists():
        print(f"ERROR: scorecard JSON not found: {in_path}", file=sys.stderr)
        return 1

    scorecard = json.loads(in_path.read_text(encoding="utf-8"))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    build(scorecard, str(out_path))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
